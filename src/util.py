from config import ratio
from openpyxl import load_workbook
from telnetlib import STATUS
from tokenize import Special
import time
from difflib import SequenceMatcher
from tabulate import tabulate
from operator import itemgetter
import re
from interfaces.person import Person

'''
Calculates how similar two strings by calling SequenceMatcher
'''
def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

'''
Analyzes file for any similar names among different states
'''
def anyalze(sortedSearches, f):
    f.write("\n\n\nSearch Matches:\n")
    hits = 0
    for i in range(0, len(sortedSearches)):
        result = []
        flag = True
        if ((not sortedSearches[i][6])):
            for x in range(i + 1, len(sortedSearches)):
                score = similar(sortedSearches[x][0].lower(), sortedSearches[i][0].lower()) #get ratio of two different names
                if ( score > ratio and (sortedSearches[x][2] != sortedSearches[i][2] or sortedSearches[x][2] == '' ) and (not sortedSearches[x][6])): #check if its not within the same state and above the ratio
                    if flag:
                        hits += 1
                        personArr = []
                        personArr.append(sortedSearches[i][0])
                        personArr.append(sortedSearches[i][1])
                        personArr.append(sortedSearches[i][2])
                        personArr.append(sortedSearches[i][3])
                        personArr.append(sortedSearches[i][4])
                        personArr.append(sortedSearches[i][5])
                        personArr.append(1)
                        sortedSearches[i][6] = True
                        result.append(personArr)
                        flag = False
                    personArr = []
                    personArr.append(sortedSearches[x][0])
                    personArr.append(sortedSearches[x][1])
                    personArr.append(sortedSearches[x][2])
                    personArr.append(sortedSearches[x][3])
                    personArr.append(sortedSearches[x][4])
                    personArr.append(sortedSearches[x][5])
                    sortedSearches[x][6] = True
                    personArr.append(round(score,2))
                    result.append(personArr)
            if (len(result) > 0):
                f.write("\n\n" + sortedSearches[i][0]+ "\n")
                sortedResults = sorted(result, key=itemgetter(6), reverse=True)
                tabulateResults = tabulate(sortedResults, headers=['Name', 'Speciality', 'State', 'Status', 'Link', 'Source', 'Similarity Score'], tablefmt='orgtbl')
                f.write(tabulateResults)
    f.write("\n\n\nTOTAL MATCHES/HITS: " + format(hits))

'''
This will create the report.txt file
'''
def createReport(searches, reportName):
    f = open(('reports/'+ reportName + '.txt'), "w")
    result = []
    for i in range(0, len(searches)): 
        personArr = []
        personArr.append(searches[i].name)
        personArr.append(searches[i].speciality)
        personArr.append(searches[i].state)
        personArr.append(searches[i].status)
        personArr.append(searches[i].link)
        personArr.append(searches[i].source)
        result.append(personArr)
    # result = sorted(result, key=itemgetter(2))
    tabulateResults = tabulate(result, headers=['Name', 'Speciality', 'State', 'Status', 'Link', 'Source'], tablefmt='orgtbl')
    f.write('MEDICAL LOOKUP REPORT\n\nFound Searches:\n')
    f.write(tabulateResults)
    for i in range(0, len(result)): 
        result[i].append(False)
    anyalze(result, f)
    f.close()
    

'''
Parses Excel sheet to get first name and last name
'''
def parseExcel(firstNameArr,lastNameArr):
    wb = load_workbook('Book2.xlsx')
    ws = wb.active
    index = 0
    for row in ws.iter_rows():
        index = 0
        for cell in row:
            if (index == 0):
                if (cell.value  != None):
                    firstNameArr.append(cell.value)
                else:
                    firstNameArr.append('')
            if (index == 1):
                if (cell.value  != None):
                    lastNameArr.append(cell.value)
                else:
                    lastNameArr.append('')
            index += 1
