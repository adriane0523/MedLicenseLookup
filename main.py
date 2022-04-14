from telnetlib import STATUS
from tokenize import Special
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from difflib import SequenceMatcher
from tabulate import tabulate
from operator import itemgetter
import re
from openpyxl import load_workbook

#TODO:
#Ask about state pratice place vs state licenses


'''
global variables initialization
'''
states = [ 'AK', 'AL', 'AR', 'AZ', 'CA', 'CO', 'CT', 'DC', 'DE', 'FL', 'GA',
           'HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD', 'ME',
           'MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH', 'NJ', 'NM',
           'NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX',
           'UT', 'VA', 'VT', 'WA', 'WI', 'WV', 'WY']
foundSearches = []
lastNameArr = []
firstNameArr = []
ratio = 0.9
f = open("report.txt", "w")
scriptFailed = False

'''
Calculates how similar two strings by calling SequenceMatcher
'''
def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

'''
Analyzes file for any similar names among different states
'''
def anyalze(sortedSearches):
    global ratio
    f.write("\n\n\nSearch Matches:\n")
    hits = 0
    for i in range(0, len(sortedSearches)):
        result = []
        flag = True
        if ((not sortedSearches[i][6])):
            for x in range(i + 1, len(sortedSearches)):
                score = similar(sortedSearches[x][0].lower(), sortedSearches[i][0].lower()) #get ratio of two different names
                if ( score > ratio and (sortedSearches[x][2] != sortedSearches[i][2] or sortedSearches[x][2] == '' ) and (not sortedSearches[x][6])): #check if its not within the same state and above the ratio
                    if flag:
                        hits += 1
                        personArr = []
                        personArr.append(sortedSearches[i][0])
                        personArr.append(sortedSearches[i][1])
                        personArr.append(sortedSearches[i][2])
                        personArr.append(sortedSearches[i][3])
                        personArr.append(sortedSearches[i][4])
                        personArr.append(sortedSearches[i][5])
                        personArr.append(1)
                        sortedSearches[i][6] = True
                        result.append(personArr)
                        flag = False
                    personArr = []
                    personArr.append(sortedSearches[x][0])
                    personArr.append(sortedSearches[x][1])
                    personArr.append(sortedSearches[x][2])
                    personArr.append(sortedSearches[x][3])
                    personArr.append(sortedSearches[x][4])
                    personArr.append(sortedSearches[x][5])
                    sortedSearches[x][6] = True
                    personArr.append(round(score,2))
                    result.append(personArr)
            if (len(result) > 0):
                f.write("\n\n" + sortedSearches[i][0]+ "\n")
                sortedResults = sorted(result, key=itemgetter(6), reverse=True)
                tabulateResults = tabulate(sortedResults, headers=['Name', 'Speciality', 'State', 'Status', 'Link', 'Source', 'Similarity Score'], tablefmt='orgtbl')
                f.write(tabulateResults)
    f.write("\n\n\nTOTAL MATCHES/HITS: " + format(hits))

'''
This will create the report.txt file
'''
def createReport(searches):
    result = []
    for i in range(0, len(searches)): 
        personArr = []
        personArr.append(searches[i].name)
        personArr.append(searches[i].speciality)
        personArr.append(searches[i].state)
        personArr.append(searches[i].status)
        personArr.append(searches[i].link)
        personArr.append(searches[i].source)
        result.append(personArr)
    # result = sorted(result, key=itemgetter(2))
    tabulateResults = tabulate(result, headers=['Name', 'Speciality', 'State', 'Status', 'Link', 'Source'], tablefmt='orgtbl')
    f.write('MEDICAL LOOKUP REPORT\n\nFound Searches:\n')
    f.write(tabulateResults)
    for i in range(0, len(result)): 
        result[i].append(False)
    anyalze(result)
    
'''
Person Object that is parsed from each website
'''
class Person:
  def __init__(self, name, link, state, status, license, source, speciality):
    self.name = name
    self.link = link
    self.state = state
    self.status = status
    self.license = license
    self.source = source
    self.speciality = speciality
    self.checked = False

'''
Parses Excel sheet to get first name and last name
'''
def parseExcel(firstNameArr,lastNameArr):
    wb = load_workbook('Book2.xlsx')
    ws = wb.active
    index = 0
    for row in ws.iter_rows():
        index = 0
        for cell in row:
            if (index == 0):
                if (cell.value  != None):
                    firstNameArr.append(cell.value)
                else:
                    firstNameArr.append('')
            if (index == 1):
                if (cell.value  != None):
                    lastNameArr.append(cell.value)
                else:
                    lastNameArr.append('')
            index += 1

#-----------------------Start of Script-----------------------
#call parse excel function
parseExcel(firstNameArr,lastNameArr)

#open chrome driver
option = webdriver.ChromeOptions()
option.Proxy = None
option.add_argument("-incognito")
option.add_experimental_option("excludeSwitches", ['enable-automation'])
# option.add_argument("--headless") 
#option.add_argument("disable-gpu")

#Creates instance of browser and links to all the websites to parse
browser = webdriver.Chrome(ChromeDriverManager().install(), options=option)
links = ["https://azbomprod.azmd.gov/glsuiteweb/clients/azbom/public/webverificationsearch.aspx",'https://apps.colorado.gov/dora/licensing/Lookup/LicenseLookup.aspx', 
        'http://profiles.ehs.state.ma.us/ProfilesV3', 'https://wybomprod.glsuite.us/GLSuiteWeb/Clients/WYBOM/Public/LicenseeSearch.aspx?SearchType=Physician', 
        'https://www.commerce.alaska.gov/cbp/main/search/professional']

'''
Below showscases all the webparsing elements of the websites. It will go through each website as if it was a user browsing it.
Once the website hits a data page it will parse that information and anayalze it for the report
'''
try:
    print('STARTING')
    print('This will take a bit...')
    #Iterates through names to get first name and last name
    for nameIndex in range(0,len(lastNameArr)):
        lastNameQuery = lastNameArr[nameIndex]
        firstNameQuery = firstNameArr[nameIndex]
        #Iterates through websites to get individual website link
        for globalIndex in range(0, len(links)):
            browser.get(links[globalIndex])
            time.sleep(3)
            #Parse AZ website
            if (globalIndex == 0):
                lastName = browser.find_element(By.XPATH,'//*[@id="ContentPlaceHolder1_txtLastName"]')
                firstName = browser.find_element(By.XPATH,'//*[@id="ContentPlaceHolder1_txtFirstName"]')
                submit = browser.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_btnName"]')
                lastName.send_keys(lastNameQuery)
                firstName.send_keys(firstNameQuery)
                submit.click()
                time.sleep(2)
                namesTable = browser.find_elements(By.XPATH,'//*[@id="ContentPlaceHolder1_dtgList"]/tbody/tr/td')
                index = 0
                name = ''
                link = ''
                count = 0
                for row in namesTable:
                    if (index > 1):
                        if (count == 3):
                            speciality = name.split('Specialty')
                            firstname = speciality[0].strip()
                            specialityname = 'N/A'
                            if (len(speciality) > 1):
                                specialityname = speciality[1].strip()
                            foundSearches.append(Person(firstname , link, 'AZ', 'N/A', '',links[globalIndex], specialityname ))
                            name = ''
                            link = ''
                            count = 0
                        if (row.text != 'Show Profile'):
                            name = row.text
                        else:
                            link = row.find_element(By.TAG_NAME, 'a')
                            link = link.get_attribute('href')
                        count += 1
                    index += 1
                if (len(namesTable) > 0):
                    speciality = name.split('Specialty')
                    foundSearches.append(Person(speciality[0].strip(), link, 'AZ', 'N/A', '',links[globalIndex],speciality[1].strip() ))
            #Parse CO website
            if (globalIndex == 1):
                dropdown = browser.find_element(By.XPATH,'//*[@id="ctl00_MainContentPlaceHolder_ucLicenseLookup_ctl03_lbMultipleCredentialTypePrefix"]')
                dropdownSelect = Select(dropdown)
                dropdownSelect.select_by_visible_text('Medical')
                lastName = browser.find_element(By.XPATH, '//*[@id="ctl00_MainContentPlaceHolder_ucLicenseLookup_ctl03_tbLastName_Contact"]')
                firstName = browser.find_element(By.XPATH, '//*[@id="ctl00_MainContentPlaceHolder_ucLicenseLookup_ctl03_tbFirstName_Contact"]')
                submit = browser.find_element(By.XPATH, '//*[@id="btnLookup"]')
                lastName.send_keys(lastNameQuery)
                firstName.send_keys(firstNameQuery)
                submit.click()
                time.sleep(2)
                header = browser.find_elements(By.XPATH,'//*[@id="ctl00_MainContentPlaceHolder_ucLicenseLookup_gvSearchResults"]/thead/tr/td/ul/li')
                indexes = []
                for i in header:
                    indexes.append(i.text)
                index = 0
                for i in indexes:    
                    if (index > 0):
                        path = '//*[@id="ctl00_MainContentPlaceHolder_ucLicenseLookup_gvSearchResults"]/thead/tr[1]/td/ul/li[' + i + ']/a'
                        browser.find_element(By.XPATH, path).click()
                        time.sleep(1)
                    namesTable = browser.find_elements(By.XPATH,'//*[@id="ctl00_MainContentPlaceHolder_ucLicenseLookup_gvSearchResults"]/tbody/tr')
                    for i in namesTable:
                        columns = i.find_elements(By.TAG_NAME, 'td')
                        foundSearches.append(Person(columns[1].text, 'N/A', columns[6].text.strip(), columns[3].text, columns[2].text,links[globalIndex], 'N/A'))
                    index += 1

                if (len(indexes) == 0 ):
                    namesTable = browser.find_elements(By.XPATH,'//*[@id="ctl00_MainContentPlaceHolder_ucLicenseLookup_gvSearchResults"]/tbody/tr')
                    for i in namesTable:
                        columns = i.find_elements(By.TAG_NAME, 'td')
                        state = 'CO'
                        if (columns[6].text.strip() == ''):
                            state = columns[6].text.strip()
                        foundSearches.append(Person(columns[1].text, 'N/A', state , columns[3].text, columns[2].text,links[globalIndex], 'N/A'))
            #Parse MA website
            if (globalIndex == 2):
                lastNameKey = browser.find_element(By.XPATH,'//*[@id="LastName"]')
                firstNameKey = browser.find_element(By.XPATH,'//*[@id="FirstName"]')
                submit = browser.find_element(By.XPATH, '/html/body/div[3]/div[1]/div/form/div[4]/div/div[1]/input')
                lastNameKey.send_keys(lastNameQuery)
                firstNameKey.send_keys(firstNameQuery)
                submit.click()
                time.sleep(1)
                try:
                    fifty = browser.find_element(By.XPATH, '/html/body/div[3]/form/div[2]/div/div[1]/div[2]/div/button[3]')
                    fifty.click()
                    time.sleep(0.2)
                except:
                    pass
                findTdTable = browser.find_elements(By.XPATH,'/html/body/div[3]/form/div[2]/div/div[2]/table/tbody')
                if (len(findTdTable) > 0):
                    tdTable = browser.find_element(By.XPATH,'/html/body/div[3]/form/div[2]/div/div[2]/table/tbody')
                    namesTable = tdTable.find_elements(By.CLASS_NAME,'tabular')
                    for i in namesTable:
                        columns = i.find_elements(By.XPATH, 'td')
                        link = i.find_elements(By.XPATH, 'td/a')
                        profileLink = 'N/A'
                        lastName =''
                        for x in link:
                            profileLink = x.get_attribute('href')
                            lastName = x.text
                        firstName1 = ''
                        try:
                            ipath = columns[1].find_element(By.XPATH, 'i')
                            firstName = ipath.text
                        except:
                            firstName = columns[1].text
                        state = 'MA'
                        if (columns[6].text.strip() == ''):
                            state = columns[6].text.strip()
                        foundSearches.append(Person(firstName + " " + lastName , profileLink,state, columns[4].text, "N/A",links[globalIndex], columns[3].text))

                try:
                    nextButton = browser.find_element(By.XPATH, '/html/body/div[3]/form/div[2]/div/div[1]/div[3]/div/button[6]')
                    while(True):
                        nextButton.click()
                        time.sleep(1)
                        tdTable = browser.find_element(By.XPATH,'/html/body/div[3]/form/div[2]/div/div[2]/table/tbody')
                        namesTable = tdTable.find_elements(By.CLASS_NAME,'tabular')
                        for i in namesTable:
                            columns = i.find_elements(By.XPATH, 'td')
                            link = i.find_elements(By.XPATH, 'td/a')
                            profileLink = 'N/A'
                            lastName =''
                            for x in link:
                                profileLink = x.get_attribute('href')
                                lastName = x.text
                            firstName = ''
                            try:
                                ipath = columns[1].find_element(By.XPATH, 'i')
                                firstName = ipath.text
                            except:
                                firstName = columns[1].text
                            state = 'MA'
                            if (columns[6].text.strip() == ''):
                                state = columns[6].text.strip()
                            foundSearches.append(Person(firstName + " " + lastName , profileLink, state , columns[4].text, "N/A",links[globalIndex], columns[3].text))
                        nextButton = browser.find_element(By.XPATH, '/html/body/div[3]/form/div[2]/div/div[1]/div[3]/div/button[6]')
                except:
                    pass
            #Parse WY website
            if globalIndex == 3:
                lastNameKey = browser.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_txtLastName"]')
                firstNameKey = browser.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_txtFirstName"]')
                submit = browser.find_element(By.XPATH, '//*[@id="ContentPlaceHolder1_btnSubmit"]')
                lastNameKey.send_keys(lastNameQuery)
                firstNameKey.send_keys(firstNameQuery)
                submit.click()
                time.sleep(1)

                tdTable = browser.find_elements(By.XPATH,'//*[@id="ContentPlaceHolder1_dtgResults"]/tbody/tr')
                index = 0
                for i in tdTable:
                    if (index > 0):
                        tdStuff = i.find_elements(By.XPATH, 'td')
                        name = tdStuff[0].text.replace(', MD', '')
                        stateIndex = 0
                        flag = True
                        state = 'WY'
                        while (stateIndex < len(states)-1 and flag):
                            if states[stateIndex] in tdStuff[1].text:
                                flag = False
                                state = states[stateIndex]
                            stateIndex += 1
                        status = 'N/A'
                        if 'Emeritus' in tdStuff[6].text:
                            status = 'Emeritus'
                        elif 'Expired' in tdStuff[6].text:
                            status = 'Expired'
                        elif 'Active' in tdStuff[6].text:
                            status = 'Active'
                        foundSearches.append(Person(name, 'N/A', state, status,"N/A",links[globalIndex], tdStuff[8].text))
                        pass
                    index += 1
            #Parse AK website
            if globalIndex == 4:
                dropdown = browser.find_element(By.XPATH,'//*[@id="ProgramId"]')
                dropdownSelect = Select(dropdown)
                dropdownSelect.select_by_visible_text('Medical')
                name = browser.find_element(By.XPATH, '//*[@id="OwnerEntityName"]')
                submit = browser.find_element(By.XPATH, '//*[@id="search"]')
                fullname = ''
                if (firstNameQuery != ''):
                    fullname = firstNameQuery + " " + lastNameQuery
                else:
                    fullname = lastNameQuery
                name.send_keys(fullname)
                submit.click()
                time.sleep(2)
                namesTable = browser.find_elements(By.XPATH, '/html/body/div/div[1]/main/article/form/div[2]/table/tbody/tr')
                for i in namesTable:
                    tdStuff = i.find_elements(By.XPATH, 'td')
                    link = tdStuff[1].find_elements(By.XPATH, 'a')
                    profileLink = ''
                    for x in link:
                        profileLink = x.get_attribute('href')
                    foundSearches.append(Person(tdStuff[3].text, 'N/A', 'AK', tdStuff[4].text ,"N/A",links[globalIndex], 'N/A'))
                nextButtonIsThere = browser.find_elements(By.XPATH, '/html/body/div/div[1]/main/article/form/div[2]/div/a[3]')
                if (len(nextButtonIsThere) > 0):
                    totalPage = browser.find_element(By.XPATH, '/html/body/div/div[1]/main/article/form/div[2]/div/a[4]')
                    total = 0
                    if (totalPage.get_attribute('data-pagenum') != None):
                        total = int(totalPage.get_attribute('data-pagenum'))
                    index = 1
                    while(index < total):
                        nextButton = browser.find_element(By.XPATH, '/html/body/div/div[1]/main/article/form/div[2]/div/a[3]')
                        nextButton.click()
                        time.sleep(5)
                        namesTable = browser.find_elements(By.XPATH, '/html/body/div/div[1]/main/article/form/div[2]/table/tbody/tr')
                        for i in namesTable:
                            tdStuff = i.find_elements(By.XPATH, 'td')
                            link = tdStuff[1].find_elements(By.XPATH, 'a')
                            profileLink = ''
                            for x in link:
                                profileLink = x.get_attribute('href')
                            foundSearches.append(Person(tdStuff[3].text, 'N/A', 'AK', tdStuff[4].text ,"N/A",links[globalIndex], 'N/A'))
                        index += 1
except Exception as e:
    print('Error Has Occured')
    print(e)
    scriptFailed = True
    browser.close()

#Calls create report function
createReport(foundSearches)

#Finish by closing file and browser
f.close()
if (not scriptFailed):
    browser.close()
    print('DONE :)')
    print('Full report written in report.txt')

